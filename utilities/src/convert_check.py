import glob
import os
import shutil
from pathlib import Path
from typing import Dict, List, Tuple

from PIL import Image

from src.preprocess_tool import FolderStruct, PreProcessTool


class PreCheckTool:
    def __init__(self, image_name = ['images','image'], label_name = ['labels','label'], mask_name = ['masks','mask']):
        self.class_list = {}
        self.folder_class = {}
        self.bbox_list = {}
        self.f = None
        self.save_path = ""
        self.image_name = image_name
        self.label_name = label_name
        self.mask_name = mask_name

    def check_image(self, path: Path) -> bool:
        """check if image is read successfully

        Arguments
        ---------
        path: path

        Returns
        -------
        bool:
            return true if image is read successfully
            else return false
        """
        import cv2
        try:
            img = cv2.imread(path)
            if img is None:
                return False
            else:
                return True
        except Exception as e:
            print(f"occurs problem {e} in {path}:")
            return False

    def check_yolo_format(self, image_path: Path, label_path: Path) -> bool:
        """check if yolo format is correct

        Arguments
        ---------
        image_path: path
        label_path: path

        Returns
        -------
        bool:
            return true if yolo format is correct
            else return false
        """
        import cv2
        img = cv2.imread(image_path)
        if img is None:
            print(f"can't read image: {image_path}")
            return False

        pixel = 1/(img.shape[0]*img.shape[1])

        with open(label_path, 'r') as file:
            lines = file.readlines()

        corrected_lines = []
        valid = True

        for line_num, line in enumerate(lines, 1):
            line = line.strip().split()
            if len(line) < 5:
                print(f"{label_path} 錯誤：在第 {line_num} 行找不到足夠的參數")
                valid = False
                continue

            x, y, w, h = map(float, line[1:5])
            if w <= 0 or h <= 0:
                print(f"{label_path} 錯誤：在第 {line_num} 行的寬度或長度為 0")
                w = max(w, pixel)
                h = max(h, pixel)
                valid = False

            corrected_lines.append(f"{line[0]} {x} {y} {w} {h}\n")

        if not valid and False:
            with open(label_path, 'w') as file:
                file.writelines(corrected_lines)

        return valid

    #下面三項基本沒用到
    def check_bbox_class_number(self, root_path: Path
        ) -> Tuple[Dict[str, List[Tuple[str, int]]], List[List[str]], Dict[str, str]]:
        """
        Arguments
        ---------
        root_path: path

        Returns
        -------
        self.folder_class: Number of instances for each class
        unique_classes: List of unique classes
        unique_img: Image paths for each class
        """
        cm = CheckMatch()
        match_list = cm.check_match(root_path)
        img_list: Dict[str, str] = {}
        unique_class: List[str] = []
        unique_classes: List[List[str]] = []
        unique_img: Dict[str, str] = {}

        for i in match_list.keys():
            path = os.path.join(i, match_list[i][1])
            img_path = os.path.join(i,match_list[i][0])
            for img_name in os.listdir(img_path):
                if not self.check_image(os.path.join(img_path,img_name)):
                    print(os.path.join(img_path,img_name) + ' load error')
                    continue
                    #return {},{},{}
                file_name, extension = os.path.splitext(img_name)
                img_list[file_name] = os.path.join(img_path,img_name)
            for j in os.listdir(path):
                file_name, extension = os.path.splitext(j)
                if extension != ".txt" or file_name not in img_list.keys():
                    continue

                if not self.check_yolo_format(img_list[file_name], os.path.join(path,j)):
                    #print(os.path.join(path,j) + ' load error')
                    continue

                with open(file=os.path.join(path,j), mode='rb') as fin:
                    data = fin.read()
                try:
                    data = data.decode('utf-8')
                except Exception as e:
                    print(e)
                    break

                lines = data.split('\n')

                for line in lines:
                    line = line.strip() # remove /n and whitspace
                    if line == '':
                        continue
                    if line is None:
                        raise ValueError('line is empty')

                    class_id, x_center, y_center, width, height, confidence = (line.split() + [0.0])[:6]
                    if class_id not in self.class_list.keys():
                        self.class_list[class_id] = 0
                    self.class_list[class_id] += 1

                    if class_id not in unique_class:
                        unique_class.append(class_id)

                unique_class.sort()
                if unique_class not in unique_classes:
                    unique_classes.append(unique_class.copy())
                    unique_img[img_list[file_name]] = os.path.join(path,j)
                unique_class.clear()

            class_list = sorted(self.class_list.items())
            self.folder_class[i] = class_list.copy()
            self.class_list.clear()

        return self.folder_class , unique_classes, unique_img

    def bbox_data(self, root_path: Path) -> List:
        """
        Arguments
        ---------
        root_path: path

        Returns
        -------
        List:
            bbox_list for each datasets
        """
        from utils.annotation_utils.alg_annotation_utils import Bbox
        cm = CheckMatch()
        match_list = cm.check_match(root_path)

        bbox_one_folder = []
        for i in match_list.keys():
            label_path = os.path.join(i, match_list[i][1])
            image_path = os.path.join(i, match_list[i][0])
            for j in os.listdir(label_path):
                if j.split(".")[-1] != "txt":
                    break

                with open(file=os.path.join(label_path,j), mode='rb') as fin:
                    data = fin.read()
                try:
                    data = data.decode('utf-8')
                except Exception as e:
                    print(e)
                    break

                lines = data.split('\n')

                for line in lines:
                    line = line.strip() # remove /n and whitspace
                    if line == '':
                        continue
                    if line is None:
                        raise ValueError('line is empty')

                    img = Image.open(os.path.join(image_path, j.replace("txt", "jpg")))
                    resolution = img.size

                    bbox = Bbox()
                    bbox.parse_bbox_from_line(line, Bbox.BboxType.XY_CENTER_WH, True, resolution[0], resolution[1])

                    class_id, x_center, y_center, width, height, confidence = (line.split() + [0.0])[:6]
                    if class_id not in self.class_list.keys():
                        self.class_list[class_id] = 0
                    self.class_list[class_id] += 1
                    #bbox_one_folder.append([bbox.xmin,bbox.ymin,bbox.xmax,bbox.ymax])
                    bbox_one_folder.append([float(x_center)-float(width)/2, float(y_center)-float(height)/2, float(x_center)+float(width)/2, float(y_center)+float(height)/2])

            self.bbox_list[i] = bbox_one_folder.copy()
            self.class_list.clear()

        return self.bbox_list

    def bbox_len(self, root_path: Path) -> List:
        """
        Arguments
        ---------
        root_path: path

        Returns
        -------
        List:
            bbox_len for each datasets
        """
        from utils.annotation_utils.alg_annotation_utils import Bbox
        limit_range = 0.01

        cm = CheckMatch()
        match_list = cm.check_match(root_path)

        bbox_one_folder = []
        for i in match_list.keys():
            label_path = os.path.join(i, match_list[i][1])
            image_path = os.path.join(i, match_list[i][0])
            for j in os.listdir(label_path):
                if j.split(".")[-1] != "txt":
                    break

                with open(file=os.path.join(label_path,j), mode='rb') as fin:
                    data = fin.read()
                try:
                    data = data.decode('utf-8')
                except Exception as e:
                    print(e)
                    break

                lines = data.split('\n')

                for line in lines:
                    line = line.strip() # remove /n and whitspace
                    if line == '':
                        continue
                    if line is None:
                        raise ValueError('line is empty')

                    img = Image.open(os.path.join(image_path, j.replace("txt", "jpg")))
                    resolution = img.size

                    bbox = Bbox()
                    bbox.parse_bbox_from_line(line, Bbox.BboxType.XY_CENTER_WH, True, resolution[0], resolution[1])

                    class_id, x_center, y_center, width, height, confidence = (line.split() + [0.0])[:6]
                    if float(width)/2 < limit_range or float(height)/2 < limit_range:
                        continue
                    if class_id not in self.class_list.keys():
                        self.class_list[class_id] = 0
                    self.class_list[class_id] += 1
                    #bbox_one_folder.append([bbox.xmin,bbox.ymin,bbox.xmax,bbox.ymax])
                    bbox_one_folder.append([float(x_center)-float(width)/2, float(y_center)-float(height)/2, float(x_center)+float(width)/2, float(y_center)+float(height)/2])

            self.bbox_list[i] = bbox_one_folder.copy()
            self.class_list.clear()

        return self.bbox_list

    def precheck_bbox_by_folder(self, root_path: Path, save_path: Path = "") -> List:
        """
        Arguments
        ---------
        root_path: path
        save_path: path

        Returns
        -------
        List:
            unique classes of this folder
        """
        cm = CheckMatch()
        match_list = cm.check_match(root_path)
        img_list = {}
        unique_class = []
        unique_classes = []
        unique_img = {}

        for i in match_list.keys():
            path = os.path.join(i, match_list[i][1])
            img_path = os.path.join(i,match_list[i][0])
            for img_name in os.listdir(img_path):
                if not self.check_image(os.path.join(img_path,img_name)):
                    print(os.path.join(img_path,img_name) + ' load error')
                    continue
                file_name, extension = os.path.splitext(img_name)
                img_list[file_name] = os.path.join(img_path,img_name)
            for j in os.listdir(path):
                file_name, extension = os.path.splitext(j)
                if extension != ".txt" or file_name not in img_list.keys():
                    continue

                if not self.check_yolo_format(img_list[file_name] ,os.path.join(path,j)):
                    #print(os.path.join(path,j) + ' load error')
                    continue

                with open(file=os.path.join(path,j), mode='rb') as fin:
                    data = fin.read()
                try:
                    data = data.decode('utf-8')
                except Exception as e:
                    print(e)
                    break

                lines = data.split('\n')

                for line in lines:
                    line = line.strip() # remove /n and whitspace
                    if line == '':
                        continue
                    if line is None:
                        raise ValueError('line is empty')

                    class_id, x_center, y_center, width, height, confidence = (line.split() + [0.0])[:6]
                    if class_id not in self.class_list.keys():
                        self.class_list[class_id] = 0
                    self.class_list[class_id] += 1

                    if class_id not in unique_class:
                        unique_class.append(class_id)

                unique_class.sort()
                if unique_class not in unique_classes:
                    unique_classes.append(unique_class.copy())
                    unique_img[img_list[file_name]] = os.path.join(path,j)
                unique_class.clear()

            class_list = sorted(self.class_list.items())
            self.folder_class[i] = class_list.copy()
            self.class_list.clear()

        if len(unique_img) == 0:
            return []

        return [int(item[0]) for item in self.folder_class[root_path]]

    def precheck_mask_by_folder(self,root_path: Path, save_path: Path) -> List:
        """
        Arguments
        ---------
        root_path: path
        save_path: path

        Returns
        -------
        List:
            unique classes of this folder
        """
        import re

        import pandas as pd

        from src.mask_to_bbox import MaskToBbox
        folder_path = os.path.dirname(root_path)

        train = pd.read_csv(os.path.join(folder_path, root_path.split(os.sep)[-1] + '_seg.csv'))
        train = train[ train['index'].notnull() ]

        unique_class = []
        max_class = []
        for i in range(0, len(train)):
            img_name = os.path.join(folder_path,os.path.normpath(train['image_path'].iloc[i]))
            lb_name = os.path.join(folder_path,os.path.normpath(train['mask_path'].iloc[i]))
            if not self.check_image(img_name) or not self.check_image(lb_name):
                print(f"{img_name} error")
                continue
            class_id = train['class_id'].iloc[i]
            if class_id not in unique_class:
                unique_class.append(class_id)

                if isinstance(class_id,str):
                    for idx in re.findall(r'\d+', class_id):
                        if int(idx) not in max_class:
                            max_class.append(int(idx))
                else:
                    if class_id not in max_class:
                        max_class.append(class_id)

        #全畫
        MaskToBbox().mask_to_bbox_by_folder(root_path,save_path, 'bbox')

        max_class.sort()
        return max_class

    def precheck_by_dataset(self, root_path: Path, save_path: Path, seg_to_obj = True) -> bool:
        """
        generate Hub (aisdatahub) and ori_and_check (dataset, check.txt, class_colors.xlsx, train.csv, test.csv)
            HUB (annosvis, bboxes, images, labels, masks, imgclass.txt)
        Arguments
        ---------
        root_path: path
            folder path
        save_path: path
            save path

        Returns
        -------
        bool: True is successfully checked
        """

        import numpy as np

        pt = PreProcessTool()
        save_path = pt.get_save_path(root_path,save_path)

        # if not os.path.exists(os.path.join(save_path, "Hub")):
        #     os.mkdir(os.path.join(save_path, "Hub"))
        # save_path = os.path.join(save_path, "ori_and_check")
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)

        csv_files = glob.glob(os.path.join(save_path,"*csv"))

        csv_file_names = ['_'.join(os.path.splitext(os.path.basename(file))[0].split('_')[:-1]) for file in csv_files]
        csv_file_names = list(np.unique(csv_file_names))
        all_unique_class = []
        task_type = ''

        for csv_file in csv_file_names:
            if not os.path.exists(os.path.join(save_path, csv_file)) or csv_file == '':
                continue
            folder_name = csv_file
            train_save_path = os.path.join(save_path,folder_name)
            # if os.path.exists(train_save_path):
            #     shutil.rmtree(train_save_path)
            # shutil.copytree(os.path.join(root_path,folder_name),train_save_path)

            if  all(name.lower() not in os.listdir(train_save_path) for name in self.mask_name):
            #    shutil.copy(os.path.join(root_path, csv_file + '_obj.csv'), save_path)
                task_type = 'obj'
                unique_class = self.precheck_bbox_by_folder(os.path.join(save_path,folder_name),train_save_path)
            else:
            #    shutil.copy(os.path.join(root_path, csv_file + '_obj.csv'), save_path)
            #    shutil.copy(os.path.join(root_path, csv_file + '_seg.csv'), save_path)
                task_type = 'seg'
                unique_class = self.precheck_mask_by_folder(os.path.join(save_path,folder_name),train_save_path)

            if unique_class == [] and task_type == 'seg':
                print(f"!!!{os.path.join(save_path,folder_name)} no labels")
                unique_class.append(0)
                continue
            all_unique_class = list(set(all_unique_class + unique_class))

        if 0 not in all_unique_class:
            all_unique_class.append(0)
        all_unique_class.sort()

        #產生check.txt
        fs = FolderStruct()
        fs.f = open(os.path.join(save_path, "check.txt"),'w+')
        self.add_transfomration_start(fs.f,'PRECHECK_BY_DATASET')
        print(f'\n{save_path}',file=fs.f)
        fs.folder_to_dict(save_path)

        #產生class_list
        print('check text written')
        print(f"class number : {len(all_unique_class)}", file=fs.f)
        if task_type == 'seg':
            print(f"class_list_seg: {all_unique_class}", file=fs.f)
            if seg_to_obj:
                all_unique_class.remove(0)
                all_unique_class = [x - 1 for x in all_unique_class]
        if task_type == 'obj' or seg_to_obj:
            print(f"class_list_obj: {all_unique_class}", file=fs.f)
        print(f"main_task_type : {task_type}", file=fs.f)
        self.add_transfomration_end(fs.f,'PRECHECK_BY_DATASET')
        fs.f.close()

        #產生顏色對應的excel
        we = WriteExcel()
        we.write_rgb_to_excel(os.path.join(save_path,'class_colors.xlsx'),all_unique_class)
        print('excel written')
        return True

    def precheck_by_folder(self, root_path: Path, save_path: Path, seg_to_obj = True) -> bool:
        """
        Arguments
        ---------
        root_path: path
            root path
        save_path: path
            save path

        Returns
        -------
        bool: True if worked well
        """
        pt = PreProcessTool()
        savepath = pt.get_save_path(root_path,save_path)
        if glob.glob(os.path.join(root_path, '*.csv')):
            print(f'check {savepath}')
            self.precheck_by_dataset(root_path, save_path, seg_to_obj)
        if os.path.isdir(savepath):
            for path in os.listdir(savepath):
                self.precheck_by_folder(os.path.join(savepath, path), save_path)
        return True

    def convert_command(self, root_path: Path, save_path: Path) -> None:
        """generate convert command

        Arguments
        ---------
        root_path: path
            root path
        save_path: path
            save path
        """
        import re

        txt_path = os.path.join(save_path,"ori_and_check")

        if self.f == None:
            self.save_path = save_path
            self.f = open(os.path.join(self.save_path, root_path.split(os.sep)[-1] + ".txt"), 'w+')

        csv_files = glob.glob(os.path.join(txt_path, '*.csv'))

        if len(csv_files) > 0:

            if os.path.exists(os.path.join(txt_path,"convert_command.txt")):
                os.remove(os.path.join(txt_path,"convert_command.txt"))
            for csv_file in csv_files:
                csv_name = '_'.join(os.path.splitext(csv_file)[0].split('_')[:-1]).split(os.sep)[-1]
                if csv_name not in os.listdir(os.path.join(save_path,"ori_and_check")):
                    continue
                csv_type = os.path.splitext(csv_file)[0].split('_')[-1]
                hub_path = os.path.join(os.path.join(save_path,"Hub"),os.path.splitext(csv_file.split(os.sep)[-1])[0])
                task_type = ''

                with open(os.path.join(txt_path,"check.txt"), 'r') as file:
                    lines = file.readlines()
                    for line in reversed(lines):
                        if line.strip().startswith('main_task_type'):
                            match = re.search(r'main_task_type\s*:\s*(.*)', line)
                            if match:
                                task_type = match.group(1).strip()
                        elif line.strip().startswith(f'class_list_{task_type}'):
                            match = re.search(r':\s*\[(.*?)\]', line)
                            if match:
                                try:
                                    last_line = list(eval(match.group(1).strip()))
                                except Exception:
                                    last_line = [eval(match.group(1).strip())]
                            break
                seg_lbs = ""
                obj_lbs = ""
                if 'seg' in csv_type:
                    for i in range(last_line[-1]):
                        lb = f'-l "{i+1}:1:class_{i}_pos" '
                        seg_lbs = seg_lbs + lb
                else:
                    if 'seg' in task_type:
                        for i in range(last_line[-1]):
                            lb =  f'-l "{i}:1:class_{i}_pos" '
                            obj_lbs = obj_lbs + lb
                    else:
                        for i in range(last_line[-1]+1):
                            lb =  f'-l "{i}:1:class_{i}_pos" '
                            obj_lbs = obj_lbs + lb

                if last_line[-1] == 0:
                    obj_lbs = f'-l "0:1:class_0_pos" '

                if os.path.exists(hub_path):
                    shutil.rmtree(hub_path)
                os.mkdir(hub_path)
                path_to_aasdatahub = shutil.which('aasdatahub')
                if seg_lbs != "":
                    write_lines = f'{path_to_aasdatahub} convert aas-dataset -r {txt_path} -n {hub_path} -m {csv_file.split(os.sep)[-1]} {seg_lbs}-ts -td\n\n'
                else:
                    write_lines = f'{path_to_aasdatahub} convert aas-dataset -r {txt_path} -n {hub_path} -m {csv_file.split(os.sep)[-1]} {obj_lbs}-td\n\n'

                self.f.writelines(write_lines)
                f = open(os.path.join(txt_path, "convert_command.txt"), 'a+')
                print(write_lines, file=f)
                f.close()

        if os.path.isdir(save_path):
            for path in os.listdir(save_path):
                self.convert_command(os.path.join(save_path,path), os.path.join(save_path,path))

    def check_command_format(self, file_path: Path) -> bool:
        """
        Arguments
        ---------
        file_path: path

        Returns
        -------
        bool:
            True if command format is correct
        """
        import re
        try:
            with open(file_path, 'r') as file:
                command = file.read().strip()
        except FileNotFoundError:
            print(f"File not found: {file_path}")
            return False
        # Define regex patterns for each part of the command
        regex_patterns = {
            'command' : r'aisdatahub convert aas-dataset',
            'r_path': r'-r "([^"]+)"',
            'n_path': r'-n "([^"]+)"',
            'm_file': r'-m "([^"]+)"',
            'labels': r'-l\s+"(.*?)"'
        }

        # Extract the paths and file names
        paths = {}
        for key, pattern in regex_patterns.items():
            if key != 'labels' and key !='command':
                match = re.search(pattern, command)
                if match:
                    paths[key] = match.group(1)

        # Validate command
        if not re.search(regex_patterns['command'], command):
            print(f"Command format is incorrect.")
            return False

        # Check if -r path exists
        if 'r_path' in paths and not os.path.exists(paths['r_path']):
            print(f"Path specified in -r does not exist: {paths['r_path']}")
            return False

        # Check if -n path exists
        if 'n_path' in paths and not os.path.exists(paths['n_path']):
            print(f"Path specified in -n does not exist: {paths['n_path']}")
            return False

        # Check if -m file exists within the -r path
        if 'm_file' in paths:
            m_file_path = os.path.join(paths['r_path'], paths['m_file'])
            if not os.path.exists(m_file_path):
                print(f"File specified in -m does not exist in -r path: {m_file_path}")
                return False

        # Check if labels format is correct
        label_matches = re.findall(regex_patterns['labels'], command)
        if not label_matches:
            print(f"No valid labels found in the command")
            return False

        # Validate each label format
        for label in label_matches:
            if not re.match(r'\d+:\d+:\w+', label):
                print(f"Invalid label format: {label}")
                return False

        return True

    def txt_run(self, root_path: Path, save_path: Path) -> None:
        """run commands generated from covert_command

        Arguments
        ---------
        root_path: path
        save_path: path
        """
        import subprocess
        pt = PreProcessTool()
        root_path = os.path.join(pt.get_save_path(root_path,save_path), root_path.split(os.sep)[-1] + ".txt")

        try:
            with open(root_path, 'r') as file:
            # read content of files and print in terminal
                for line in file:
                    line = line.strip()
                    # generate terminal messages
                    command = f"{line}"
                    command = command.split(' ')
                    if 'aasdatahub' not in command[0]:
                        continue
                    # 在終端機中執行命令並捕捉輸出
                    result = subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=False)
                    # print the result
                    print(result.stdout.decode())
        except subprocess.CalledProcessError as e:
            print("occurs problem while running the messages:", e)
        except FileNotFoundError:
            print("can't find file:", root_path)
        except Exception as e:
            print("occurs unexpected problem:", e)

    def check_hub(self, root_path: Path, save_path: Path) -> List[str]:
        """check datahub's records

        Arguments
        ---------
        root_path: path
            root path
        save_path: path
            save path

        Returns
        -------
        List[str]: printed_str
            to store the ouputs, for testing
        """
        import re
        import subprocess

        from aasdatahub import ImageHub
        from aasdatahub.structures import ImageTask
        pt = PreProcessTool()
        save_path = pt.get_save_path(root_path,save_path)
        printed_str = []
        for item in os.listdir(save_path):
            folder_path = os.path.join(save_path,item)
            if os.path.isdir(folder_path):
                if 'hub' == item.lower():
                    check_path = os.path.join(save_path,'ori_and_check')
                    with open(os.path.join(check_path,'check.txt'),'r+') as check_file:
                        lines = check_file.readlines()
                        for line in reversed(lines):
                            if line.strip().startswith(f'class_list_seg'):
                                match = re.search(r':\s*\[(.*?)\]', line)
                                if match:
                                    seg_last_line = eval(match.group(1).strip())
                                    if not isinstance(seg_last_line,tuple):
                                        seg_last_line = [seg_last_line]
                            if line.strip().startswith(f'class_list_obj'):
                                match = re.search(r':\s*\[(.*?)\]', line)
                                if match:
                                    bbox_last_line = eval(match.group(1).strip())
                                    if not isinstance(bbox_last_line,tuple):
                                        bbox_last_line = [bbox_last_line]

                        self.add_transfomration_start(check_file,'CHECK_HUB')
                        status = True
                        for hub_folder in os.listdir(folder_path):
                            hub = ImageHub()
                            mask_loaded = False
                            bbox_loaded = False
                            try:
                                hub.load(os.path.join(folder_path,hub_folder))
                            except Exception:
                                printed_str.append(f'{os.path.join(folder_path,hub_folder)}: load failed')
                                print(f'{os.path.join(folder_path,hub_folder)}: load failed')
                                continue
                            printed_str.append(f"\n{hub_folder}:")
                            print(f"\n{hub_folder}:",file=check_file)

                            try:
                                mask_loaded = hub.read_mask(0).is_labeled
                                printed_str.append(f"mask_loaded: {mask_loaded}")
                                print(f"mask_loaded: {mask_loaded}",file=check_file)
                            except Exception:
                                printed_str.append(f"mask_loaded: False")
                                print(f"mask_loaded: False",file=check_file)
                                continue

                            try:
                                bbox_loaded = hub.read_bboxes(0).is_labeled
                                printed_str.append(f"bbox_loaded: {bbox_loaded}")
                                print(f"bbox_loaded: {bbox_loaded}",file=check_file)
                            except Exception:
                                printed_str.append(f"bbox_loaded: False")
                                print(f"bbox_loaded: False",file=check_file)
                                continue

                            if mask_loaded:
                                printed_str.append(f"class_name_seg: {dict(zip(hub.get_classname_table(ImageTask.SEGMENTATION), seg_last_line))}")
                                print(f"class_name_seg: {dict(zip(hub.get_classname_table(ImageTask.SEGMENTATION), seg_last_line))}",file=check_file)
                            if bbox_loaded:
                                printed_str.append(f"class_name_bbox: {dict(zip(hub.get_classname_table(ImageTask.OBJECT_DETECTION), bbox_last_line))}")
                                print(f"class_name_bbox: {dict(zip(hub.get_classname_table(ImageTask.OBJECT_DETECTION), bbox_last_line))}",file=check_file)

                            if not mask_loaded and not bbox_loaded:
                                status = False
                            else:
                                try:
                                    command = ["python", "-m", "src.utils.fifityone.app","--dataset-path", os.path.join(folder_path,hub_folder),"--destination-dir", './src/utils/fifityone/artifacts',"--overwrite", "--persistent"]
                                    result = subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=False)
                                    print(result.stdout.decode())
                                    pass
                                except Exception as e:
                                    pass

                        if status:
                            printed_str.append(f"status: successfully")
                            print(f"\nstatus: successfully",file=check_file)
                        else:
                            printed_str.append(f"status: fail")
                            print(f"\nstatus: fail",file=check_file)
                        self.add_transfomration_end(check_file,'CHECK_HUB')
                        check_file.close()
                else:
                    if os.path.exists(folder_path):
                        self.check_hub(folder_path,save_path)
        return printed_str

    def add_transfomration_start(self, file, name: str) -> None:
        """
        write=======================================================\n>>>>{name}_START.

        Arguments
        ---------
        file:
            already been opened txt file
        name: str
            name of the function
        """
        transformation_line = f"=======================================================\n>>>>{name}_START.\n"

        try:
            # Open the file in append mode and add the transformation line
            file.write(transformation_line)
        except FileNotFoundError:
            print(f"Error: The file at {file.name} was not found.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def add_transfomration_end(self, file, name: str) -> None:
        """
        write\n=======================================================\n>>>>{name}_END.

        Arguments
        ---------
        file:
            had been opened txt file
        name:
            name of the function
        """
        transformation_line = f"\n=======================================================\n>>>>{name}_END.\n\n"

        try:
            # Open the file in append mode and add the transformation line
            file.write(transformation_line)
        except FileNotFoundError:
            print(f"Error: The file at {file.name} was not found.")
        except Exception as e:
            print(f"An error occurred: {e}")


    def checking_all(self, root_path: Path, save_path: Path, check_res = True, seg_to_obj=True) -> None:
        """to check every funtion

        Arguments
        ---------
        root_path: path
            root path
        save_path: path
            save path
        check_res: bool
            is set to True, by default
        """
        pt = PreProcessTool()
        self.precheck_by_folder(root_path, save_path, seg_to_obj)
        save_path = pt.get_save_path(root_path,save_path)
        self.convert_command(root_path, save_path)
        file_name = self.f.name
        self.f.close()
        if not check_res:
            print(f'\nYou need to modify the class names at this path:\n{file_name}\n')
            input('After completing the modification, press enter to continue...')
            check_res = self.check_command_format(os.path.join(pt.get_save_path(root_path,save_path), root_path.split(os.sep)[-1] + ".txt"))
        while True:
            if check_res:
                print('run command')
                self.txt_run(root_path, save_path)
                print('check hub')
                self.check_hub(root_path, save_path)
                break
            else:
                print('Format error detected. Please revise and correct the input format.\n')
                input('After completing the modification, press enter to continue...')
                check_res = self.check_command_format(os.path.join(pt.get_save_path(root_path,save_path), root_path.split(os.sep)[-1] + ".txt"))

class WriteExcel:

  def write_rgb_to_excel(self, excel_file: Path, unique_class: list) -> None:
    """
    Arguments
    ---------
    excel_file: path
        path of the excel file
    unique class: list
        list of different classes
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    rgb_data = [
    [0, 0, 0],
    [128, 0, 0],
    [0, 128, 0],
    [128, 128, 0],
    [0, 0, 128],
    [128, 0, 128],
    [0, 128, 128],
    [128, 128, 128],
    [64, 0, 0],
    [192, 0, 0],
    [64, 128, 0],
    [192, 128, 0],
    [64, 0, 128],
    [192, 0, 128],
    [64, 128, 128],
    [192, 128, 128],
    [0, 64, 0],
    [128, 64, 0],
    [0, 192, 0],
    [128, 192, 0],
    [0, 64, 128]
    ]
    wb = Workbook()
    ws = wb.active
    cc = 0
    col = 3
    for i, rgb in enumerate(rgb_data):
        fill = PatternFill(start_color=f'FF{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}',
                            end_color = f'FF{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}', fill_type = 'solid')
        ws.cell(row = i + 1, column = 2, value = f'{rgb}')
        ws.cell(row = i + 1, column = 1).fill = fill
        if i in unique_class:
            if i > len(rgb_data):
                col += 1
            ws.cell(row=i+2, column=col, value=unique_class[cc])
            cc += 1
    if len(unique_class) > len(rgb_data):
        for i in range(len(rgb_data),len(unique_class)):
            ws.cell(row = i + 1, column = col, value = unique_class[i])

    wb.save(excel_file)

class CheckMatch:

    def __init__(self, image_name = ['images','image'], label_name = ['labels','label'], mask_name = ['masks','mask']) -> None:
        self.match_path = ""
        self.match_list = {}
        self.mask_list = {}
        self.txt_list = {}
        self.img_list = {}
        self.palette_list = {}
        self.parent_name = ""
        self.supported_file_keys = []
        self.image_name = image_name
        self.label_name = label_name
        self.mask_name = mask_name
        self.pc = PreCheckTool()

    def check_match_loop(self, path: Path, folder_structure: FolderStruct) -> List:
        """
        Arguments
        ---------
        path: path
            the path of the folder
        folder_structure:
            the path's structure

        Returns
        -------
        list
            return a list, with image and label's path and name inside
        """
        pt = PreProcessTool()
        if self.parent_name != folder_structure.parent_name:
            self.mask_list = {}
            self.txt_list = {}
            self.img_list = {}
            self.palette_list = {}
            self.parent_name = folder_structure.parent_name
        masks_num = 0
        image_num = 0
        match_path = os.path.join(path, folder_structure.folder_name)
        if folder_structure.repo_dict:
            for repo in folder_structure.repo_dict:
                self.check_match_loop(match_path, folder_structure.repo_dict[repo])
            if self.txt_list and self.mask_list and self.img_list and self.palette_list:
                for tvalue in self.mask_list.keys():
                    if tvalue in self.img_list.keys() and tvalue in self.txt_list.keys() and tvalue in self.palette_list.keys():
                        self.match_list[self.match_path] = [self.img_list[tvalue], self.txt_list[tvalue], self.mask_list[tvalue], self.palette_list[tvalue]]
                        self.palette_list.clear()
                        self.mask_list.clear()
                        self.img_list.clear()
                        self.txt_list.clear()
                        break
            if self.txt_list and self.mask_list and self.img_list:
                for tvalue in self.mask_list.keys():
                    if tvalue in self.img_list.keys() and tvalue in self.txt_list.keys():
                        self.match_list[self.match_path] = [self.img_list[tvalue], self.txt_list[tvalue], self.mask_list[tvalue], f'{self.mask_name[0]}_palette']
                        self.mask_list.clear()
                        self.img_list.clear()
                        self.txt_list.clear()
                        break
            if self.txt_list and self.img_list:
                for tvalue in self.txt_list.keys():
                    if tvalue in self.img_list.keys():
                        self.match_list[self.match_path] = [self.img_list[tvalue], self.txt_list[tvalue]]
                        self.txt_list.clear()
                        self.img_list.clear()
                        break
            if self.mask_list and self.img_list:
                for tvalue in self.mask_list.keys():
                    if tvalue in self.img_list.keys():
                        self.match_list[self.match_path] = [self.img_list[tvalue], self.mask_list[tvalue], f'{self.mask_name[0]}_palette']
                        self.mask_list.clear()
                        self.img_list.clear()
                        break

        else:
            if ".txt" in folder_structure.file_dict.keys():
                self.txt_list[folder_structure.file_dict[".txt"]] = folder_structure.folder_name
            if folder_structure.img_dict:
                for extension in folder_structure.img_dict.keys():
                    if extension[0] in self.supported_file_keys:
                        if any(name.lower() in folder_structure.folder_name.lower() for name in self.image_name):
                            self.img_list[str(image_num + int(folder_structure.img_dict[extension][0]))] = folder_structure.folder_name
                            self.img_list.pop(image_num,f'無{image_num}')
                            image_num += int(folder_structure.img_dict[extension][0])
                            self.match_path = path
                        elif '_palette' in folder_structure.folder_name.lower():
                            self.palette_list[folder_structure.img_dict[extension][0]] = folder_structure.folder_name
                        elif any(name.lower() in folder_structure.folder_name.lower() for name in self.mask_name):
                            self.mask_list[str(masks_num + int(folder_structure.img_dict[extension][0]))] = folder_structure.folder_name
                            self.mask_list.pop(masks_num,f'無{image_num}')
                            masks_num += int(folder_structure.img_dict[extension][0])
                            if f'{folder_structure.folder_name}_palette' not in os.listdir(os.path.dirname(match_path)):
                                if pt.is_rgb_folder(match_path):
                                    pt.color_to_gray(match_path)
                                else:
                                    pt.gray_to_color(match_path)

        return self.match_list.copy()

    def check_match(self, path: Path) -> List:
        """
        Argguments
        ----------
        path: path

        Returns
        -------
        List
            return a List of pair, with image and label's path and name inside
        """
        fs = FolderStruct()
        folder_structure = fs.folder_to_dict(path)
        self.parent_name = folder_structure.parent_name
        self.supported_file_keys = fs.SUPPORTED_FILE_KEYS
        match_list = self.check_match_loop(os.path.dirname(path), folder_structure)
        self.__init__()

        return match_list

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("-p","--path", type=str, default=r"D:\data\temp\customer\U.D.ELECTRONIC", help="your path")
    parser.add_argument("-sp","--savepath", type=str, default=r'D:\data\test', help="your save path")
    parser.add_argument("-m","--mode", type=str, default="checkhub", help="which to do")
    args = parser.parse_args()

    pc = PreCheckTool()
    pt = PreProcessTool()

    if args.mode == "checkmatch":
        cm = CheckMatch()
        match_list = cm.check_match(args.path)
        for key, value in match_list.items():
            print(f"{key}: {value}")

    if args.mode == "checkclass":

        folder_class = pc.check_bbox_class_number(args.path)
        for item in folder_class.items():
            print(str(item[0]) + ": " + str(len(item[1])))

    if args.mode == "bboxlist":
        sp = args.savepath + os.sep + str(len(os.listdir(args.savepath))+1) + "-" + args.path.split(os.sep)[-1] + "-bboxlist"
        f = open(sp + ".txt","w+")
        result = pc.bbox_data(args.path)
        for folder in result.keys():
            for bbox in result[folder]:
                if abs(bbox[0]-bbox[2]) < 0.001 or abs(bbox[1]-bbox[3]) < 0.001 :
                    pass
                print(f"{bbox}", file=f)
        f.close()

    if args.mode == "precheck":
        pc.precheck_by_folder(args.path, args.savepath)

    if args.mode =="aisdatahub":
        pc.convert_command(args.path, args.savepath)
        pc.f.close()

    if args.mode == "runcmd":
        pc.txt_run(args.path, args.savepath)

    if args.mode == 'checkhub':
        pc.check_hub(args.path, args.savepath)

    if args.mode == "checking":
        '''
        path:       convert folder
        save path:  datahub folder
        '''
        pc.checking_all(args.path, args.savepath)

    print('end')
